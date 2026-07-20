"""
لوحة بيانات ويب لمراقبة أداء بوت التداول — خدمة مستقلة تماماً عن البوت
نفسه، تتصل بنفس قاعدة البيانات (DATABASE_URL) للقراءة فقط (لا تُعدّل أي
شيء في البوت أو صفقاته إطلاقاً — read-only بالكامل، آمنة تماماً).
"""
import io
import os
import datetime
import logging

import asyncpg
from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("dashboard")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
FALLBACK_DATABASE_URL = os.getenv("FALLBACK_DATABASE_URL", "").strip()

app = FastAPI(title="لوحة بيانات بوت التداول")

_pool: asyncpg.Pool = None


async def get_pool() -> asyncpg.Pool:
    """يُنشئ مجمع اتصال واحداً فقط (Primary أولاً، Fallback عند فشله)."""
    global _pool
    if _pool is not None:
        return _pool
    try:
        _pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=3, command_timeout=15)
        logger.info("✅ اتصال لوحة البيانات بقاعدة البيانات الأساسية نجح")
    except Exception as e:
        logger.warning(f"فشل الاتصال بالقاعدة الأساسية ({e})، تجربة الاحتياطية...")
        _pool = await asyncpg.create_pool(FALLBACK_DATABASE_URL, min_size=1, max_size=3, command_timeout=15)
        logger.info("✅ اتصال لوحة البيانات بقاعدة البيانات الاحتياطية نجح")
    return _pool


def _parse_date_range(date_from: str = None, date_to: str = None):
    """يحوّل تواريخ نصية (YYYY-MM-DD) إلى نطاق timestamp، بحدود افتراضية معقولة إن غابت."""
    now = datetime.datetime.now(datetime.timezone.utc)
    if date_to:
        end_dt = datetime.datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=datetime.timezone.utc)
        end_ts = (end_dt + datetime.timedelta(days=1)).timestamp()  # نهاية اليوم المحدد بالكامل
    else:
        end_ts = now.timestamp()

    if date_from:
        start_dt = datetime.datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=datetime.timezone.utc)
        start_ts = start_dt.timestamp()
    else:
        start_ts = 0  # بلا حد أدنى — كل التاريخ

    return start_ts, end_ts


@app.get("/api/stats")
async def api_stats(date_from: str = Query(None), date_to: str = Query(None)):
    """إحصائيات مُلخَّصة: عدد الصفقات المفتوحة/المغلقة، نسبة الربح، صافي الربح/الخسارة، الرصيد الحالي."""
    pool = await get_pool()
    start_ts, end_ts = _parse_date_range(date_from, date_to)

    open_row = await pool.fetchrow("SELECT COUNT(*) as c FROM trades WHERE status = 'open'")

    closed_row = await pool.fetchrow("""
        SELECT
            COUNT(*) as total_closed,
            SUM(CASE WHEN profit_loss_sol >= 0 THEN 1 ELSE 0 END) as winning,
            SUM(CASE WHEN profit_loss_sol < 0 THEN 1 ELSE 0 END) as losing,
            COALESCE(SUM(profit_loss_sol), 0) as net_pl,
            COALESCE(SUM(capital_invested_sol), 0) as total_deployed
        FROM trades
        WHERE status IN ('closed_profit', 'closed_loss', 'closed_flagged')
          AND exit_timestamp >= $1 AND exit_timestamp < $2
    """, start_ts, end_ts)

    total_closed = closed_row["total_closed"] or 0
    winning = closed_row["winning"] or 0
    win_rate = (winning / total_closed * 100) if total_closed else 0.0

    # الرصيد الحالي: يُقرأ من آخر صفقة (أياً كانت فتحاً أو إغلاقاً) لأننا
    # نُسجّل الرصيد الفعلي بعد كل عملية بالفعل ضمن رسائل التنبيه — هنا
    # نستنتجه من قاعدة البيانات مباشرة عبر مجموع (رأس المال الأصلي +
    # صافي الربح/الخسارة التراكمي)، وهو تقريب معقول بدون استدعاء RPC حي.
    lifetime_row = await pool.fetchrow("""
        SELECT COALESCE(SUM(profit_loss_sol), 0) as lifetime_pl
        FROM trades WHERE status IN ('closed_profit', 'closed_loss', 'closed_flagged')
    """)

    return JSONResponse({
        "open_trades": open_row["c"] or 0,
        "closed_trades": total_closed,
        "winning_trades": winning,
        "losing_trades": closed_row["losing"] or 0,
        "win_rate_pct": round(win_rate, 1),
        "net_profit_loss_sol": round(closed_row["net_pl"] or 0, 4),
        "total_capital_deployed_sol": round(closed_row["total_deployed"] or 0, 4),
        "lifetime_profit_loss_sol": round(lifetime_row["lifetime_pl"] or 0, 4),
    })


@app.get("/api/trades")
async def api_trades(date_from: str = Query(None), date_to: str = Query(None), status: str = Query("all")):
    """قائمة الصفقات ضمن نطاق تاريخي، مع فلتر اختياري للحالة (مفتوحة/مغلقة/الكل)."""
    pool = await get_pool()
    start_ts, end_ts = _parse_date_range(date_from, date_to)

    status_filter = ""
    if status == "open":
        status_filter = "AND status = 'open'"
    elif status == "closed":
        status_filter = "AND status IN ('closed_profit', 'closed_loss', 'closed_flagged')"

    rows = await pool.fetch(f"""
        SELECT id, mint_address, symbol, entry_timestamp, exit_timestamp,
               capital_invested_sol, proceeds_sol, profit_loss_sol, status,
               close_reason, strategy, tx_hash_entry, tx_hash_exit
        FROM trades
        WHERE (
            (entry_timestamp >= $1 AND entry_timestamp < $2)
            OR (exit_timestamp >= $1 AND exit_timestamp < $2)
        )
        {status_filter}
        ORDER BY entry_timestamp DESC
        LIMIT 1000
    """, start_ts, end_ts)

    trades = []
    for r in rows:
        capital = r["capital_invested_sol"] or 0
        pl_sol = r["profit_loss_sol"]
        pl_pct = (pl_sol / capital * 100) if (pl_sol is not None and capital) else None
        trades.append({
            "id": r["id"],
            "symbol": r["symbol"],
            "mint_address": r["mint_address"],
            "entry_time": datetime.datetime.fromtimestamp(r["entry_timestamp"], tz=datetime.timezone.utc).isoformat() if r["entry_timestamp"] else None,
            "exit_time": datetime.datetime.fromtimestamp(r["exit_timestamp"], tz=datetime.timezone.utc).isoformat() if r["exit_timestamp"] else None,
            "capital_sol": capital,
            "proceeds_sol": r["proceeds_sol"],
            "profit_loss_sol": pl_sol,
            "profit_loss_pct": round(pl_pct, 2) if pl_pct is not None else None,
            "status": r["status"],
            "close_reason": r["close_reason"],
            "strategy": r["strategy"] or "momentum_chase",
        })

    return JSONResponse({"trades": trades})


@app.get("/api/export")
async def api_export(date_from: str = Query(None), date_to: str = Query(None)):
    """يُصدّر كل صفقات النطاق الزمني المحدد كملف إكسل جاهز للتحميل المباشر."""
    pool = await get_pool()
    start_ts, end_ts = _parse_date_range(date_from, date_to)

    rows = await pool.fetch("""
        SELECT symbol, mint_address, entry_timestamp, exit_timestamp,
               capital_invested_sol, proceeds_sol, profit_loss_sol, status,
               close_reason, strategy, tx_hash_entry, tx_hash_exit
        FROM trades
        WHERE (entry_timestamp >= $1 AND entry_timestamp < $2)
        ORDER BY entry_timestamp DESC
    """, start_ts, end_ts)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الصفقات"
    ws.sheet_view.rightToLeft = True

    headers = [
        "الرمز", "عنوان العملة", "وقت الفتح", "وقت الإغلاق", "رأس المال (SOL)",
        "العائد (SOL)", "الربح/الخسارة (SOL)", "الربح/الخسارة (%)", "الحالة",
        "سبب الإغلاق", "الاستراتيجية", "معاملة الشراء", "معاملة البيع",
    ]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    profit_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    loss_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, r in enumerate(rows, start=2):
        capital = r["capital_invested_sol"] or 0
        pl_sol = r["profit_loss_sol"]
        pl_pct = (pl_sol / capital * 100) if (pl_sol is not None and capital) else None
        entry_dt = datetime.datetime.fromtimestamp(r["entry_timestamp"], tz=datetime.timezone.utc) if r["entry_timestamp"] else None
        exit_dt = datetime.datetime.fromtimestamp(r["exit_timestamp"], tz=datetime.timezone.utc) if r["exit_timestamp"] else None

        values = [
            r["symbol"], r["mint_address"],
            entry_dt.strftime("%Y-%m-%d %H:%M:%S") if entry_dt else "",
            exit_dt.strftime("%Y-%m-%d %H:%M:%S") if exit_dt else "",
            round(capital, 4),
            round(r["proceeds_sol"], 4) if r["proceeds_sol"] is not None else "",
            round(pl_sol, 4) if pl_sol is not None else "",
            round(pl_pct, 2) if pl_pct is not None else "",
            r["status"], r["close_reason"] or "", r["strategy"] or "momentum_chase",
            r["tx_hash_entry"] or "", r["tx_hash_exit"] or "",
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            if col_idx == 7 and pl_sol is not None:  # عمود الربح/الخسارة
                cell.fill = profit_fill if pl_sol >= 0 else loss_fill

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(14, len(h) + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"trades_export_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/", response_class=HTMLResponse)
async def dashboard_page():
    return HTML_PAGE


HTML_PAGE = open(os.path.join(os.path.dirname(__file__), "index.html"), "r", encoding="utf-8").read()
