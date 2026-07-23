"""
لوحة بيانات ويب لمراقبة أداء بوت التداول — خدمة مستقلة تماماً عن البوت
نفسه، تتصل بنفس قاعدة البيانات (DATABASE_URL) للقراءة فقط (لا تُعدّل أي
شيء في البوت أو صفقاته إطلاقاً — read-only بالكامل، آمنة تماماً).
"""
import io
import os
import datetime
import logging

import asyncpg
from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("dashboard")

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
FALLBACK_DATABASE_URL = os.getenv("FALLBACK_DATABASE_URL", "").strip()

app = FastAPI(title="لوحة بيانات بوت التداول")

_pool: asyncpg.Pool = None


async def get_pool() -> asyncpg.Pool:
    """يُنشئ مجمع اتصال واحداً فقط (Primary أولاً، Fallback عند فشله)."""
    global _pool
    if _pool is not None:
        return _pool
    try:
        _pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=3, command_timeout=15)
        logger.info("✅ اتصال لوحة البيانات بقاعدة البيانات الأساسية نجح")
    except Exception as e:
        logger.warning(f"فشل الاتصال بالقاعدة الأساسية ({e})، تجربة الاحتياطية...")
        _pool = await asyncpg.create_pool(FALLBACK_DATABASE_URL, min_size=1, max_size=3, command_timeout=15)
        logger.info("✅ اتصال لوحة البيانات بقاعدة البيانات الاحتياطية نجح")
    return _pool


def _parse_date_range(date_from: str = None, date_to: str = None):
    """يحوّل تواريخ نصية (YYYY-MM-DD) إلى نطاق timestamp، بحدود افتراضية معقولة إن غابت."""
    now = datetime.datetime.now(datetime.timezone.utc)
    if date_to:
        end_dt = datetime.datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=datetime.timezone.utc)
        end_ts = (end_dt + datetime.timedelta(days=1)).timestamp()  # نهاية اليوم المحدد بالكامل
    else:
        end_ts = now.timestamp()

    if date_from:
        start_dt = datetime.datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=datetime.timezone.utc)
        start_ts = start_dt.timestamp()
    else:
        start_ts = 0  # بلا حد أدنى — كل التاريخ

    return start_ts, end_ts


@app.get("/api/stats")
async def api_stats(date_from: str = Query(None), date_to: str = Query(None)):
    """إحصائيات مُلخَّصة: عدد الصفقات المفتوحة/المغلقة، نسبة الربح، صافي الربح/الخسارة، الرصيد الحالي."""
    pool = await get_pool()
    start_ts, end_ts = _parse_date_range(date_from, date_to)

    open_row = await pool.fetchrow("SELECT COUNT(*) as c FROM trades WHERE status = 'open'")

    closed_row = await pool.fetchrow("""
        SELECT
            COUNT(*) as total_closed,
            SUM(CASE WHEN profit_loss_sol >= 0 THEN 1 ELSE 0 END) as winning,
            SUM(CASE WHEN profit_loss_sol < 0 THEN 1 ELSE 0 END) as losing,
            COALESCE(SUM(profit_loss_sol), 0) as net_pl,
            COALESCE(SUM(capital_invested_sol), 0) as total_deployed
        FROM trades
        WHERE status IN ('closed_profit', 'closed_loss', 'closed_flagged')
          AND exit_timestamp >= $1 AND exit_timestamp < $2
    """, start_ts, end_ts)

    total_closed = closed_row["total_closed"] or 0
    winning = closed_row["winning"] or 0
    win_rate = (winning / total_closed * 100) if total_closed else 0.0

    # الرصيد الحالي: يُقرأ من آخر صفقة (أياً كانت فتحاً أو إغلاقاً) لأننا
    # نُسجّل الرصيد الفعلي بعد كل عملية بالفعل ضمن رسائل التنبيه — هنا
    # نستنتجه من قاعدة البيانات مباشرة عبر مجموع (رأس المال الأصلي +
    # صافي الربح/الخسارة التراكمي)، وهو تقريب معقول بدون استدعاء RPC حي.
    lifetime_row = await pool.fetchrow("""
        SELECT COALESCE(SUM(profit_loss_sol), 0) as lifetime_pl
        FROM trades WHERE status IN ('closed_profit', 'closed_loss', 'closed_flagged')
    """)

    return JSONResponse({
        "open_trades": open_row["c"] or 0,
        "closed_trades": total_closed,
        "winning_trades": winning,
        "losing_trades": closed_row["losing"] or 0,
        "win_rate_pct": round(win_rate, 1),
        "net_profit_loss_sol": round(closed_row["net_pl"] or 0, 4),
        "total_capital_deployed_sol": round(closed_row["total_deployed"] or 0, 4),
        "lifetime_profit_loss_sol": round(lifetime_row["lifetime_pl"] or 0, 4),
    })


@app.get("/api/trades")
async def api_trades(date_from: str = Query(None), date_to: str = Query(None), status: str = Query("all")):
    """قائمة الصفقات ضمن نطاق تاريخي، مع فلتر اختياري للحالة (مفتوحة/مغلقة/الكل)."""
    pool = await get_pool()
    start_ts, end_ts = _parse_date_range(date_from, date_to)

    status_filter = ""
    if status == "open":
        status_filter = "AND status = 'open'"
    elif status == "closed":
        status_filter = "AND status IN ('closed_profit', 'closed_loss', 'closed_flagged')"

    rows = await pool.fetch(f"""
        SELECT id, mint_address, symbol, entry_timestamp, exit_timestamp,
               capital_invested_sol, proceeds_sol, profit_loss_sol, status,
               close_reason, strategy, tx_hash_entry, tx_hash_exit
        FROM trades
        WHERE (
            (entry_timestamp >= $1 AND entry_timestamp < $2)
            OR (exit_timestamp >= $1 AND exit_timestamp < $2)
        )
        {status_filter}
        ORDER BY entry_timestamp DESC
        LIMIT 1000
    """, start_ts, end_ts)

    trades = []
    for r in rows:
        capital = r["capital_invested_sol"] or 0
        pl_sol = r["profit_loss_sol"]
        pl_pct = (pl_sol / capital * 100) if (pl_sol is not None and capital) else None
        trades.append({
            "id": r["id"],
            "symbol": r["symbol"],
            "mint_address": r["mint_address"],
            "entry_time": datetime.datetime.fromtimestamp(r["entry_timestamp"], tz=datetime.timezone.utc).isoformat() if r["entry_timestamp"] else None,
            "exit_time": datetime.datetime.fromtimestamp(r["exit_timestamp"], tz=datetime.timezone.utc).isoformat() if r["exit_timestamp"] else None,
            "capital_sol": capital,
            "proceeds_sol": r["proceeds_sol"],
            "profit_loss_sol": pl_sol,
            "profit_loss_pct": round(pl_pct, 2) if pl_pct is not None else None,
            "status": r["status"],
            "close_reason": r["close_reason"],
            "strategy": r["strategy"] or "momentum_chase",
        })

    return JSONResponse({"trades": trades})


@app.get("/api/export")
async def api_export(date_from: str = Query(None), date_to: str = Query(None)):
    """يُصدّر كل صفقات النطاق الزمني المحدد كملف إكسل جاهز للتحميل المباشر."""
    pool = await get_pool()
    start_ts, end_ts = _parse_date_range(date_from, date_to)

    rows = await pool.fetch("""
        SELECT symbol, mint_address, entry_timestamp, exit_timestamp,
               capital_invested_sol, proceeds_sol, profit_loss_sol, status,
               close_reason, strategy, tx_hash_entry, tx_hash_exit
        FROM trades
        WHERE (entry_timestamp >= $1 AND entry_timestamp < $2)
        ORDER BY entry_timestamp DESC
    """, start_ts, end_ts)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الصفقات"
    ws.sheet_view.rightToLeft = True

    headers = [
        "الرمز", "عنوان العملة", "وقت الفتح", "وقت الإغلاق", "رأس المال (SOL)",
        "العائد (SOL)", "الربح/الخسارة (SOL)", "الربح/الخسارة (%)", "الحالة",
        "سبب الإغلاق", "الاستراتيجية", "معاملة الشراء", "معاملة البيع",
    ]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    profit_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    loss_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, r in enumerate(rows, start=2):
        capital = r["capital_invested_sol"] or 0
        pl_sol = r["profit_loss_sol"]
        pl_pct = (pl_sol / capital * 100) if (pl_sol is not None and capital) else None
        entry_dt = datetime.datetime.fromtimestamp(r["entry_timestamp"], tz=datetime.timezone.utc) if r["entry_timestamp"] else None
        exit_dt = datetime.datetime.fromtimestamp(r["exit_timestamp"], tz=datetime.timezone.utc) if r["exit_timestamp"] else None

        values = [
            r["symbol"], r["mint_address"],
            entry_dt.strftime("%Y-%m-%d %H:%M:%S") if entry_dt else "",
            exit_dt.strftime("%Y-%m-%d %H:%M:%S") if exit_dt else "",
            round(capital, 4),
            round(r["proceeds_sol"], 4) if r["proceeds_sol"] is not None else "",
            round(pl_sol, 4) if pl_sol is not None else "",
            round(pl_pct, 2) if pl_pct is not None else "",
            r["status"], r["close_reason"] or "", r["strategy"] or "momentum_chase",
            r["tx_hash_entry"] or "", r["tx_hash_exit"] or "",
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            if col_idx == 7 and pl_sol is not None:  # عمود الربح/الخسارة
                cell.fill = profit_fill if pl_sol >= 0 else loss_fill

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(14, len(h) + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"trades_export_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/", response_class=HTMLResponse)
async def dashboard_page():
    return HTML_PAGE


HTML_PAGE = """<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>لوحة أداء البوت</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;700&family=IBM+Plex+Sans+Arabic:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  :root {
    --bg: #0A0E17;
    --surface: #131826;
    --surface-2: #1A2032;
    --border: #232B3D;
    --text: #EDEFF4;
    --text-dim: #8B93A7;
    --amber: #E8A33D;
    --amber-dim: #7A5A26;
    --green: #3ECF8E;
    --red: #F0625A;
    --radius: 10px;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0;
    background: var(--bg);
    color: var(--text);
    font-family: 'IBM Plex Sans Arabic', sans-serif;
    min-height: 100vh;
  }
  .mono { font-family: 'JetBrains Mono', monospace; font-variant-numeric: tabular-nums; }

  header {
    padding: 28px 32px 20px;
    border-bottom: 1px solid var(--border);
    display: flex;
    align-items: baseline;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 12px;
  }
  header h1 {
    font-size: 20px;
    font-weight: 600;
    margin: 0;
    display: flex;
    align-items: center;
    gap: 10px;
  }
  .pulse-dot {
    width: 8px; height: 8px; border-radius: 50%;
    background: var(--green);
    box-shadow: 0 0 0 0 rgba(62,207,142,0.6);
    animation: pulse 2s infinite;
  }
  @keyframes pulse {
    0% { box-shadow: 0 0 0 0 rgba(62,207,142,0.5); }
    70% { box-shadow: 0 0 0 8px rgba(62,207,142,0); }
    100% { box-shadow: 0 0 0 0 rgba(62,207,142,0); }
  }
  header .last-update { color: var(--text-dim); font-size: 13px; }

  main { padding: 28px 32px 60px; max-width: 1280px; margin: 0 auto; }

  .stats-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 14px;
    margin-bottom: 28px;
  }
  .stat-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 18px 20px;
    position: relative;
    overflow: hidden;
  }
  .stat-card.highlight {
    border-color: var(--amber-dim);
    background: linear-gradient(160deg, var(--surface) 0%, #1C1508 140%);
  }
  .stat-label {
    font-size: 12px;
    color: var(--text-dim);
    margin-bottom: 8px;
    letter-spacing: 0.02em;
  }
  .stat-value {
    font-size: 26px;
    font-weight: 700;
    line-height: 1.1;
  }
  .stat-value.amber { color: var(--amber); }
  .stat-value.pos { color: var(--green); }
  .stat-value.neg { color: var(--red); }
  .stat-sub { font-size: 12px; color: var(--text-dim); margin-top: 6px; }

  .controls {
    display: flex;
    align-items: center;
    gap: 10px;
    flex-wrap: wrap;
    margin-bottom: 18px;
    padding: 14px 16px;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
  }
  .controls label { font-size: 13px; color: var(--text-dim); }
  .controls input[type=date] {
    background: var(--surface-2);
    border: 1px solid var(--border);
    color: var(--text);
    padding: 7px 10px;
    border-radius: 6px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px;
  }
  .btn {
    background: var(--surface-2);
    border: 1px solid var(--border);
    color: var(--text);
    padding: 8px 16px;
    border-radius: 6px;
    font-size: 13px;
    cursor: pointer;
    font-family: inherit;
    transition: all 0.15s;
  }
  .btn:hover { border-color: var(--amber-dim); }
  .btn.primary {
    background: var(--amber);
    color: #1A1200;
    border-color: var(--amber);
    font-weight: 600;
  }
  .btn.primary:hover { filter: brightness(1.1); }
  .btn-group { display: flex; gap: 6px; margin-inline-start: auto; }
  .btn.small { padding: 6px 12px; font-size: 12px; }
  .btn.active { background: var(--amber-dim); border-color: var(--amber); }

  table {
    width: 100%;
    border-collapse: collapse;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    font-size: 13px;
  }
  thead th {
    background: var(--surface-2);
    color: var(--text-dim);
    font-weight: 500;
    text-align: right;
    padding: 10px 14px;
    border-bottom: 1px solid var(--border);
    font-size: 12px;
    white-space: nowrap;
  }
  tbody td {
    padding: 10px 14px;
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
  }
  tbody tr:last-child td { border-bottom: none; }
  tbody tr:hover { background: var(--surface-2); }
  .pl-pos { color: var(--green); }
  .pl-neg { color: var(--red); }
  .badge {
    display: inline-block;
    padding: 2px 9px;
    border-radius: 20px;
    font-size: 11px;
    background: var(--surface-2);
    border: 1px solid var(--border);
    color: var(--text-dim);
  }
  .badge.open { color: var(--amber); border-color: var(--amber-dim); }
  .strategy-tag {
    font-size: 11px;
    color: var(--text-dim);
    font-family: 'JetBrains Mono', monospace;
  }
  .empty-state {
    text-align: center;
    padding: 50px 20px;
    color: var(--text-dim);
  }
  .table-wrap { overflow-x: auto; }
  footer { text-align: center; padding: 30px; color: var(--text-dim); font-size: 12px; }
</style>
</head>
<body>

<header>
  <h1><span class="pulse-dot"></span> لوحة أداء البوت</h1>
  <div class="last-update mono" id="lastUpdate">—</div>
</header>

<main>
  <div class="stats-grid">
    <div class="stat-card highlight">
      <div class="stat-label">صافي الربح/الخسارة (الفترة المحددة)</div>
      <div class="stat-value mono amber" id="netPL">—</div>
    </div>
    <div class="stat-card">
      <div class="stat-label">صافي الربح/الخسارة (كل الوقت)</div>
      <div class="stat-value mono" id="lifetimePL">—</div>
    </div>
    <div class="stat-card">
      <div class="stat-label">صفقات مفتوحة الآن</div>
      <div class="stat-value mono" id="openTrades">—</div>
    </div>
    <div class="stat-card">
      <div class="stat-label">صفقات مغلقة (الفترة)</div>
      <div class="stat-value mono" id="closedTrades">—</div>
      <div class="stat-sub" id="closedSub">—</div>
    </div>
    <div class="stat-card">
      <div class="stat-label">نسبة الربح</div>
      <div class="stat-value mono" id="winRate">—</div>
    </div>
  </div>

  <div class="controls">
    <label>من</label>
    <input type="date" id="dateFrom">
    <label>إلى</label>
    <input type="date" id="dateTo">
    <button class="btn" onclick="applyFilter()">تطبيق</button>
    <div class="btn-group">
      <button class="btn small" onclick="quickRange(1)">اليوم</button>
      <button class="btn small" onclick="quickRange(7)">7 أيام</button>
      <button class="btn small" onclick="quickRange(30)">30 يوماً</button>
      <button class="btn small" onclick="quickRange(0)">كل الوقت</button>
      <button class="btn primary small" onclick="exportExcel()">⬇ تحميل Excel</button>
    </div>
  </div>

  <div class="controls" style="padding: 8px 16px;">
    <button class="btn small active" id="filterAll" onclick="setStatusFilter('all')">الكل</button>
    <button class="btn small" id="filterOpen" onclick="setStatusFilter('open')">مفتوحة فقط</button>
    <button class="btn small" id="filterClosed" onclick="setStatusFilter('closed')">مغلقة فقط</button>
  </div>

  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>الرمز</th>
          <th>الاستراتيجية</th>
          <th>وقت الفتح</th>
          <th>وقت الإغلاق</th>
          <th>رأس المال</th>
          <th>الربح/الخسارة</th>
          <th>النسبة</th>
          <th>الحالة</th>
        </tr>
      </thead>
      <tbody id="tradesBody">
        <tr><td colspan="8" class="empty-state">جارٍ التحميل...</td></tr>
      </tbody>
    </table>
  </div>
</main>

<footer>لوحة قراءة فقط — لا تُعدّل أي بيانات في البوت</footer>

<script>
let statusFilter = 'all';

function fmtDate(iso) {
  if (!iso) return '—';
  const d = new Date(iso);
  return d.toLocaleString('ar-SA', { year: 'numeric', month: '2-digit', day: '2-digit', hour: '2-digit', minute: '2-digit' });
}

function qs() {
  const from = document.getElementById('dateFrom').value;
  const to = document.getElementById('dateTo').value;
  let params = [];
  if (from) params.push('date_from=' + from);
  if (to) params.push('date_to=' + to);
  return params.length ? '?' + params.join('&') : '';
}

function quickRange(days) {
  const to = new Date();
  document.getElementById('dateTo').value = to.toISOString().slice(0, 10);
  if (days === 0) {
    document.getElementById('dateFrom').value = '';
  } else {
    const from = new Date();
    from.setDate(from.getDate() - days);
    document.getElementById('dateFrom').value = from.toISOString().slice(0, 10);
  }
  applyFilter();
}

function setStatusFilter(s) {
  statusFilter = s;
  ['All', 'Open', 'Closed'].forEach(x => document.getElementById('filter' + x).classList.remove('active'));
  document.getElementById('filter' + s.charAt(0).toUpperCase() + s.slice(1)).classList.add('active');
  loadTrades();
}

async function applyFilter() {
  await Promise.all([loadStats(), loadTrades()]);
}

async function loadStats() {
  const res = await fetch('/api/stats' + qs());
  const d = await res.json();

  const netEl = document.getElementById('netPL');
  netEl.textContent = (d.net_profit_loss_sol >= 0 ? '+' : '') + d.net_profit_loss_sol.toFixed(4) + ' SOL';
  netEl.className = 'stat-value mono ' + (d.net_profit_loss_sol >= 0 ? 'pos' : 'neg');

  const lifeEl = document.getElementById('lifetimePL');
  lifeEl.textContent = (d.lifetime_profit_loss_sol >= 0 ? '+' : '') + d.lifetime_profit_loss_sol.toFixed(4) + ' SOL';
  lifeEl.className = 'stat-value mono ' + (d.lifetime_profit_loss_sol >= 0 ? 'pos' : 'neg');

  document.getElementById('openTrades').textContent = d.open_trades;
  document.getElementById('closedTrades').textContent = d.closed_trades;
  document.getElementById('closedSub').textContent = d.winning_trades + ' رابحة / ' + d.losing_trades + ' خاسرة';
  document.getElementById('winRate').textContent = d.win_rate_pct + '%';

  document.getElementById('lastUpdate').textContent = 'آخر تحديث: ' + new Date().toLocaleTimeString('ar-SA');
}

async function loadTrades() {
  const res = await fetch('/api/trades' + qs() + (qs() ? '&' : '?') + 'status=' + statusFilter);
  const d = await res.json();
  const body = document.getElementById('tradesBody');

  if (!d.trades.length) {
    body.innerHTML = '<tr><td colspan="8" class="empty-state">لا توجد صفقات ضمن هذه الفترة</td></tr>';
    return;
  }

  body.innerHTML = d.trades.map(t => {
    const isOpen = t.status === 'open';
    const plClass = t.profit_loss_sol == null ? '' : (t.profit_loss_sol >= 0 ? 'pl-pos' : 'pl-neg');
    const plText = t.profit_loss_sol == null ? '—' : (t.profit_loss_sol >= 0 ? '+' : '') + t.profit_loss_sol.toFixed(4);
    const pctText = t.profit_loss_pct == null ? '—' : (t.profit_loss_pct >= 0 ? '+' : '') + t.profit_loss_pct.toFixed(1) + '%';
    const statusBadge = isOpen
      ? '<span class="badge open">مفتوحة</span>'
      : '<span class="badge">' + (t.close_reason ? t.close_reason.slice(0, 30) : 'مغلقة') + '</span>';

    return `<tr>
      <td><strong>${t.symbol || '?'}</strong></td>
      <td><span class="strategy-tag">${t.strategy}</span></td>
      <td class="mono">${fmtDate(t.entry_time)}</td>
      <td class="mono">${fmtDate(t.exit_time)}</td>
      <td class="mono">${t.capital_sol ? t.capital_sol.toFixed(4) : '—'}</td>
      <td class="mono ${plClass}">${plText}</td>
      <td class="mono ${plClass}">${pctText}</td>
      <td>${statusBadge}</td>
    </tr>`;
  }).join('');
}

function exportExcel() {
  window.location.href = '/api/export' + qs();
}

applyFilter();
setInterval(applyFilter, 30000); // تحديث تلقائي كل 30 ثانية
</script>

</body>
</html>
"""
